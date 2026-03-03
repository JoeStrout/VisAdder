#!/usr/bin/env python3
"""
Generate tinyadder_4.xlsx — a spreadsheet implementing the full forward pass
of the 4-digit TinyAdder transformer using spreadsheet formulas.

Edit the token strings in row 2 and watch all intermediate computations update.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter
from math import log, exp

# === Constants (mirroring tinyadder_4.py) ===
NUM_DIGITS = 4
NUM_CANDIDATES = 10
TOKENS = [str(i) for i in range(10)] + ["=", "<bos>", "<eos>", "+"]
DIGIT_EMBED_SCALE = 10
V_SCALE = 1e4
DIGIT_SCALE = 10 ** NUM_DIGITS
FINAL_SCALE = 100
DIGIT_OFFSET = 0.5
GATE_BIAS_SHIFT = 15.0
ALIBI_CONSTANT = log(10)
K_DIGIT_SCORE = -1000.0
K_SPECIAL_SCORE = -40.0
V_PROJ_SPECIAL = 0.1
V_PROJ_NEG_DOUBLE = -1.1
V_PROJ_SCALE = exp(K_SPECIAL_SCORE - log(10))

EQ_DIM, SPECIAL_DIM, DIGIT_DIM, COUNT_DIM, SCALE_DIM = 0, 1, 2, 3, 4

# Spreadsheet layout: columns B..Q = positions 0..15 (max 16 tokens)
MAX_POS = 16
COL_OFFSET = 2  # column B = 2

# Embedding table: 14 rows x 5 cols
EMB_TABLE = []
for tok_id in range(14):
    row = [0.0] * 5
    if 1 <= tok_id <= 9:
        row[DIGIT_DIM] = tok_id * DIGIT_EMBED_SCALE
    if tok_id == 10:  # "="
        row[EQ_DIM] = 1.0
        row[SPECIAL_DIM] = 1.0
    if tok_id == 11:  # "<bos>"
        row[SPECIAL_DIM] = 1.0
    if tok_id == 13:  # "+"
        row[SPECIAL_DIM] = 1.0
    EMB_TABLE.append(row)

# FFN0 up_vals
UP_VALS = [(d + DIGIT_OFFSET) * DIGIT_SCALE * FINAL_SCALE for d in range(NUM_CANDIDATES)]
UP_VALS.append(DIGIT_SCALE)

DIM_NAMES_5 = ["EQ", "SPECIAL", "DIGIT", "COUNT", "SCALE"]
DIM_NAMES_16 = DIM_NAMES_5 + [f"cand{d}" for d in range(NUM_CANDIDATES)] + ["digit_pos"]

# Styles
HEADER_FONT = Font(bold=True, size=12)
SECTION_FILL = PatternFill("solid", fgColor="4472C4")
SECTION_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBSECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
SUBSECTION_FONT = Font(bold=True)
DIM_FONT = Font(italic=True, color="666666")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")


def col(pos):
    """Column letter for token position pos (0-based)."""
    return get_column_letter(COL_OFFSET + pos)


def cell(pos, row):
    """Cell reference like B5 for position pos, row number row."""
    return f"{col(pos)}{row}"


def col_range(row, start_pos=0, end_pos=MAX_POS - 1):
    """Range like B5:Q5."""
    return f"{cell(start_pos, row)}:{cell(end_pos, row)}"


class SheetBuilder:
    """Builds the main computation sheet row by row."""

    def __init__(self, ws):
        self.ws = ws
        self.row = 1  # current row pointer

    def current_row(self):
        return self.row

    def section_header(self, title):
        """Write a section header spanning all columns."""
        r = self.row
        self.ws.cell(r, 1, title).font = SECTION_FONT
        self.ws.cell(r, 1).fill = SECTION_FILL
        for p in range(MAX_POS):
            self.ws.cell(r, COL_OFFSET + p).fill = SECTION_FILL
        self.ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COL_OFFSET + MAX_POS - 1)
        self.row += 1
        return r

    def subsection_header(self, title):
        """Write a subsection header."""
        r = self.row
        self.ws.cell(r, 1, title).font = SUBSECTION_FONT
        self.ws.cell(r, 1).fill = SUBSECTION_FILL
        for p in range(MAX_POS):
            self.ws.cell(r, COL_OFFSET + p).fill = SUBSECTION_FILL
        self.row += 1
        return r

    def label_row(self, label):
        """Write a dim label in column A, return the row number, advance."""
        r = self.row
        c = self.ws.cell(r, 1, label)
        c.font = DIM_FONT
        self.row += 1
        return r

    def skip(self, n=1):
        self.row += n

    def write_cell(self, row, pos, value):
        """Write a value to a cell at (row, position)."""
        self.ws.cell(row, COL_OFFSET + pos, value)

    def write_formula(self, row, pos, formula):
        """Write a formula to a cell at (row, position)."""
        self.ws.cell(row, COL_OFFSET + pos, formula)

    def set_format(self, row, pos, fmt):
        self.ws.cell(row, COL_OFFSET + pos).number_format = fmt


def build_ref_sheet(wb):
    """Create the Ref sheet with embedding table, token list, and constants."""
    ws = wb.create_sheet("Ref")

    # --- Token list (A1:A14) ---
    ws.cell(1, 1, "Token").font = Font(bold=True)
    ws.cell(1, 2, "ID").font = Font(bold=True)
    for i, tok in enumerate(TOKENS):
        ws.cell(i + 2, 1, tok)
        ws.cell(i + 2, 2, i)
    # Named range: TokenList = Ref!A2:A15
    # We'll reference it as Ref!$A$2:$A$15

    # --- Embedding table (D1:H14, with header) ---
    ws.cell(1, 4, "EmbTable").font = Font(bold=True)
    for d, name in enumerate(DIM_NAMES_5):
        ws.cell(1, 4 + d, name).font = Font(bold=True)
    for tok_id in range(14):
        for d in range(5):
            ws.cell(tok_id + 2, 4 + d, EMB_TABLE[tok_id][d])
    # EmbTable = Ref!$D$2:$H$15

    # --- FFN0 up_vals (J1:J12) ---
    ws.cell(1, 10, "up_vals").font = Font(bold=True)
    for i, v in enumerate(UP_VALS):
        ws.cell(i + 2, 10, v)
    # up_vals = Ref!$J$2:$J$12

    # --- Constants (L column) ---
    ws.cell(1, 12, "Constant").font = Font(bold=True)
    ws.cell(1, 13, "Value").font = Font(bold=True)
    constants = [
        ("DIGIT_SCALE", DIGIT_SCALE),
        ("FINAL_SCALE", FINAL_SCALE),
        ("V_SCALE", V_SCALE),
        ("GATE_BIAS_SHIFT", GATE_BIAS_SHIFT),
        ("K_DIGIT_SCORE", K_DIGIT_SCORE),
        ("K_SPECIAL_SCORE", K_SPECIAL_SCORE),
        ("V_PROJ_SPECIAL", V_PROJ_SPECIAL),
        ("V_PROJ_NEG_DOUBLE", V_PROJ_NEG_DOUBLE),
        ("V_PROJ_SCALE", V_PROJ_SCALE),
        ("ALIBI_CONSTANT", ALIBI_CONSTANT),
        ("K_WEIGHT", K_SPECIAL_SCORE - K_DIGIT_SCORE),
        ("V_W1", V_PROJ_SPECIAL / V_PROJ_SCALE),
        ("V_W2", V_PROJ_NEG_DOUBLE / V_PROJ_SCALE),
    ]
    for i, (name, val) in enumerate(constants):
        ws.cell(i + 2, 12, name)
        ws.cell(i + 2, 13, val)

    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['L'].width = 20
    ws.column_dimensions['M'].width = 25

    return ws


def build_input_section(sb, default_tokens):
    """Build the input section: token strings (editable) and token IDs."""
    sb.section_header("INPUT")

    # Row: Token strings (user-editable)
    tok_row = sb.label_row("Token")
    for p, tok in enumerate(default_tokens):
        sb.write_cell(tok_row, p, tok)
        sb.ws.cell(tok_row, COL_OFFSET + p).fill = INPUT_FILL

    # Row: Token IDs (formula lookup)
    id_row = sb.label_row("Token ID")
    for p in range(MAX_POS):
        c = cell(p, tok_row)
        f = f'=IF({c}="","",MATCH({c},Ref!$A$2:$A$15,0)-1)'
        sb.write_formula(id_row, p, f)

    # Seq length
    sb.ws.cell(id_row, COL_OFFSET + MAX_POS + 1, "Seq Length:")
    sb.ws.cell(id_row, COL_OFFSET + MAX_POS + 2,
               f'=COUNTA({cell(0, tok_row)}:{cell(MAX_POS-1, tok_row)})')

    return tok_row, id_row


def build_embedding_section(sb, id_row):
    """Build the embedding section: 5 dims x 16 positions."""
    sb.section_header("EMBEDDING (5 dims)")
    emb_rows = []
    for d in range(5):
        r = sb.label_row(DIM_NAMES_5[d])
        emb_rows.append(r)
        for p in range(MAX_POS):
            id_cell = cell(p, id_row)
            # INDEX into Ref!$D$2:$H$15 (14 rows x 5 cols)
            f = f'=IF({id_cell}="",0,INDEX(Ref!$D$2:$H$15,{id_cell}+1,{d+1}))'
            sb.write_formula(r, p, f)
    return emb_rows


def build_l0_head3(sb, emb_rows):
    """Build L0 Attention Head 3 (ALiBi counter).
    Returns the output row."""
    sb.section_header("L0 ATTENTION HEAD 3 (ALiBi counter)")

    # K[j] = SPECIAL[j] * (K_SPECIAL - K_DIGIT) + K_DIGIT
    # = SPECIAL[j] * 960 + (-1000)
    sb.subsection_header("K (key)")
    k_row = sb.label_row("K")
    special_row = emb_rows[SPECIAL_DIM]
    for p in range(MAX_POS):
        sp = cell(p, special_row)
        f = f'={sp}*960+(-1000)'
        sb.write_formula(k_row, p, f)

    # V[j] = SPECIAL[j] * V_W1 + EQ[j] * V_W2
    sb.subsection_header("V (value)")
    v_row = sb.label_row("V")
    eq_row = emb_rows[EQ_DIM]
    v_w1 = V_PROJ_SPECIAL / V_PROJ_SCALE
    v_w2 = V_PROJ_NEG_DOUBLE / V_PROJ_SCALE
    for p in range(MAX_POS):
        sp = cell(p, special_row)
        eq = cell(p, eq_row)
        f = f'={sp}*{v_w1}+{eq}*{v_w2}'
        sb.write_formula(v_row, p, f)

    # Raw scores: Q=1, so score[i,j] = K[j]
    # ALiBi[i,j] = (j - i) * ln(10)
    # Causal mask: j > i → -1000
    sb.subsection_header("Scores + ALiBi + Causal Mask (16x16)")
    score_rows = []
    for i in range(MAX_POS):
        r = sb.label_row(f"i={i}")
        score_rows.append(r)
        for j in range(MAX_POS):
            k_cell = cell(j, k_row)
            alibi = (j - i) * ALIBI_CONSTANT
            if j > i:
                sb.write_cell(r, j, -1000)
            else:
                f = f'={k_cell}+{alibi}'
                sb.write_formula(r, j, f)

    # Softmax1 over each row
    sb.subsection_header("Softmax1 (16x16)")
    sm_rows = []
    for i in range(MAX_POS):
        r = sb.label_row(f"i={i}")
        sm_rows.append(r)
        for j in range(MAX_POS):
            score_cell = cell(j, score_rows[i])
            # sum of EXP for the row (all positions)
            exp_sum_parts = []
            for jj in range(MAX_POS):
                exp_sum_parts.append(f"EXP({cell(jj, score_rows[i])})")
            exp_sum = "+".join(exp_sum_parts)
            f = f'=EXP({score_cell})/(1+{exp_sum})'
            sb.write_formula(r, j, f)
            sb.set_format(r, j, '0.000000')

    # Output[i] = SUMPRODUCT(softmax1_row_i, V_row)
    sb.subsection_header("Head 3 Output")
    h3_out_row = sb.label_row("out")
    for i in range(MAX_POS):
        sm_range = col_range(sm_rows[i])
        v_range = col_range(v_row)
        f = f'=SUMPRODUCT({sm_range},{v_range})'
        sb.write_formula(h3_out_row, i, f)

    return h3_out_row


def build_l0_head4(sb, emb_rows):
    """Build L0 Attention Head 4 (gate/scale signal).
    K = -1000 everywhere, Q = 1, no ALiBi, V = EQ[j] * 1.0
    Returns the output row."""
    sb.section_header("L0 ATTENTION HEAD 4 (gate signal)")

    # V[j] = EQ[j]
    sb.subsection_header("V (value)")
    v4_row = sb.label_row("V")
    eq_row = emb_rows[EQ_DIM]
    for p in range(MAX_POS):
        f = f'={cell(p, eq_row)}'
        sb.write_formula(v4_row, p, f)

    # Scores: Q*K = 1 * -1000 = -1000 always; causal mask: j>i → -1000 (same)
    # So all visible scores = -1000, future = -1000
    # softmax1(-1000) for all = exp(-1000)/(1 + n*exp(-1000)) ≈ 0 for non-eq
    # But we need to show it properly. Since K=-1000 for all, scores are all -1000.
    # softmax1: exp(-1000)/(1 + sum(exp(-1000))) for each visible position
    # This is essentially uniform but tiny. The output picks up EQ contribution.

    # Actually let's think more carefully. All scores = -1000, so
    # exp(-1000) ≈ 0, softmax1 ≈ 0 for each position, output ≈ 0.
    # But that can't be right because head 4 produces the scale signal...

    # Wait - looking at the code more carefully, in the actual model:
    # k[..., SCALE_HEAD] is NOT set (stays 0). Only ADJUSTMENT_HEAD gets k values.
    # So for head 4 (SCALE_HEAD=4), K=0 for all positions.
    # With K=0, Q=1: score = Q*K = 0 for all visible positions.
    # Causal mask: j>i → -inf. Visible: score=0.
    # softmax1(0) = exp(0)/(1 + n*exp(0)) = 1/(1+n) where n = number of visible positions.
    # This gives uniform causal attention with softmax1!

    sb.subsection_header("Scores (all 0, causal masked)")
    score4_rows = []
    for i in range(MAX_POS):
        r = sb.label_row(f"i={i}")
        score4_rows.append(r)
        for j in range(MAX_POS):
            if j > i:
                sb.write_cell(r, j, -1000)
            else:
                sb.write_cell(r, j, 0)

    sb.subsection_header("Softmax1 (16x16)")
    sm4_rows = []
    for i in range(MAX_POS):
        r = sb.label_row(f"i={i}")
        sm4_rows.append(r)
        for j in range(MAX_POS):
            score_cell = cell(j, score4_rows[i])
            exp_sum_parts = [f"EXP({cell(jj, score4_rows[i])})" for jj in range(MAX_POS)]
            exp_sum = "+".join(exp_sum_parts)
            f = f'=EXP({score_cell})/(1+{exp_sum})'
            sb.write_formula(r, j, f)
            sb.set_format(r, j, '0.000000')

    sb.subsection_header("Head 4 Output")
    h4_out_row = sb.label_row("out")
    for i in range(MAX_POS):
        sm_range = col_range(sm4_rows[i])
        v_range = col_range(v4_row)
        f = f'=SUMPRODUCT({sm_range},{v_range})'
        sb.write_formula(h4_out_row, i, f)

    return h4_out_row


def build_l0_attn_residual(sb, emb_rows, h3_out_row, h4_out_row):
    """Combine heads into attention output, add residual to embedding."""
    sb.section_header("L0 ATTENTION OUTPUT + RESIDUAL")

    # Attention output: [0, 0, 0, head3, head4] per position
    # head3 → dim 3 (COUNT), head4 → dim 4 (SCALE)
    sb.subsection_header("Attn Output (5 dims)")
    attn_rows = []
    for d in range(5):
        r = sb.label_row(DIM_NAMES_5[d])
        attn_rows.append(r)
        for p in range(MAX_POS):
            if d == 3:  # COUNT ← head3
                f = f'={cell(p, h3_out_row)}'
            elif d == 4:  # SCALE ← head4
                f = f'={cell(p, h4_out_row)}'
            else:
                sb.write_cell(r, p, 0)
                continue
            sb.write_formula(r, p, f)

    # Residual: embedding + attn output
    sb.subsection_header("Residual = Embedding + Attn Output")
    resid_rows = []
    for d in range(5):
        r = sb.label_row(DIM_NAMES_5[d])
        resid_rows.append(r)
        for p in range(MAX_POS):
            emb_c = cell(p, emb_rows[d])
            attn_c = cell(p, attn_rows[d])
            f = f'={emb_c}+{attn_c}'
            sb.write_formula(r, p, f)

    return resid_rows


def build_l0_ffn(sb, resid_rows):
    """Build L0 FFN: gate * up for 11 dims."""
    sb.section_header("L0 FFN (gated, 11 dims)")

    # Gate[d=0..9] = MAX(0, SCALE_DIM)
    # Gate[10] = MAX(0, DIGIT_DIM)
    sb.subsection_header("Gate (ReLU)")
    gate_rows = []
    for d in range(NUM_CANDIDATES + 1):
        label = f"gate[{d}]" if d < NUM_CANDIDATES else "gate[10]"
        r = sb.label_row(label)
        gate_rows.append(r)
        for p in range(MAX_POS):
            if d < NUM_CANDIDATES:
                src = cell(p, resid_rows[SCALE_DIM])
            else:
                src = cell(p, resid_rows[DIGIT_DIM])
            f = f'=MAX(0,{src})'
            sb.write_formula(r, p, f)

    # Up[d] = COUNT_DIM * up_val[d]
    sb.subsection_header("Up = COUNT * up_val")
    up_rows = []
    for d in range(NUM_CANDIDATES + 1):
        label = f"up[{d}]"
        r = sb.label_row(label)
        up_rows.append(r)
        for p in range(MAX_POS):
            count_c = cell(p, resid_rows[COUNT_DIM])
            # Reference up_val from Ref sheet
            f = f'={count_c}*Ref!$J${d+2}'
            sb.write_formula(r, p, f)

    # Output = gate * up
    sb.subsection_header("FFN Output = Gate * Up")
    ffn_rows = []
    for d in range(NUM_CANDIDATES + 1):
        label = f"ffn[{d}]"
        r = sb.label_row(label)
        ffn_rows.append(r)
        for p in range(MAX_POS):
            g = cell(p, gate_rows[d])
            u = cell(p, up_rows[d])
            f = f'={g}*{u}'
            sb.write_formula(r, p, f)

    return ffn_rows


def build_widened_residual(sb, resid_rows, ffn_rows):
    """Build widened residual: 5 original dims + 11 FFN dims = 16 dims."""
    sb.section_header("WIDENED RESIDUAL (16 dims)")

    wide_rows = []
    # First 5 dims: just copy from resid
    for d in range(5):
        r = sb.label_row(DIM_NAMES_16[d])
        wide_rows.append(r)
        for p in range(MAX_POS):
            f = f'={cell(p, resid_rows[d])}'
            sb.write_formula(r, p, f)

    # Dims 5..15: candidates (5..14) and digit_pos (15)
    # These come from FFN output added to the widened slot (which was 0)
    for d in range(NUM_CANDIDATES + 1):
        r = sb.label_row(DIM_NAMES_16[5 + d])
        wide_rows.append(r)
        for p in range(MAX_POS):
            f = f'={cell(p, ffn_rows[d])}'
            sb.write_formula(r, p, f)

    return wide_rows


def build_l1_attention(sb, wide_rows):
    """Build L1 Attention: Q=K=0, uniform causal softmax1.
    V = digit_pos_dim * 100 + 15."""
    sb.section_header("L1 ATTENTION (uniform causal softmax1)")

    # V[j] = wide_rows[DIGIT_POS_DIM index] * FINAL_SCALE + GATE_BIAS_SHIFT
    # DIGIT_POS_DIM is index 15 in wide_rows
    digit_pos_idx = 15  # index in wide_rows for digit_pos dim
    sb.subsection_header("V = digit_pos * 100 + 15")
    v1_row = sb.label_row("V")
    for p in range(MAX_POS):
        dp = cell(p, wide_rows[digit_pos_idx])
        f = f'={dp}*100+15'
        sb.write_formula(v1_row, p, f)

    # Softmax1 with all scores=0: weight = 1/(i+2) for each visible position
    # (i+1 visible positions, plus 1 for softmax1 denominator)
    sb.subsection_header("Softmax1 weights (uniform causal)")
    sm1_rows = []
    for i in range(MAX_POS):
        r = sb.label_row(f"i={i}")
        sm1_rows.append(r)
        for j in range(MAX_POS):
            if j > i:
                sb.write_cell(r, j, 0)
            else:
                # 1/(1 + (i+1)) = 1/(i+2)
                f = f'=1/({i}+2)'
                sb.write_formula(r, j, f)
            sb.set_format(r, j, '0.000000')

    # Output[i] = SUMPRODUCT(softmax1_row_i, V_row)
    sb.subsection_header("L1 Attn Output")
    l1_out_row = sb.label_row("out")
    for i in range(MAX_POS):
        sm_range = col_range(sm1_rows[i])
        v_range = col_range(v1_row)
        f = f'=SUMPRODUCT({sm_range},{v_range})'
        sb.write_formula(l1_out_row, i, f)

    return l1_out_row


def build_l1_residual(sb, wide_rows, l1_out_row):
    """L1 residual: widened + attn output (broadcast to all dims)."""
    # Actually L1 attn output is 1 dim, broadcast-added to all 16 dims?
    # No — looking at the code: out is shape (B, S, 1), then h = h + attn1_out
    # This broadcasts: adds the single value to ALL 16 dims.
    sb.section_header("L1 RESIDUAL (16 dims)")
    l1_resid_rows = []
    for d in range(16):
        r = sb.label_row(DIM_NAMES_16[d])
        l1_resid_rows.append(r)
        for p in range(MAX_POS):
            w = cell(p, wide_rows[d])
            a = cell(p, l1_out_row)
            f = f'={w}+{a}'
            sb.write_formula(r, p, f)
    return l1_resid_rows


def build_candidates_and_ffn1(sb, l1_resid_rows):
    """Extract candidates (dims 5-14), apply L1 FFN (|x| via dual ReLU)."""
    sb.section_header("CANDIDATES (dims 5-14)")
    cand_rows = []
    for d in range(NUM_CANDIDATES):
        r = sb.label_row(f"cand[{d}]")
        cand_rows.append(r)
        for p in range(MAX_POS):
            # dims 5..14 in l1_resid_rows
            f = f'={cell(p, l1_resid_rows[5 + d])}'
            sb.write_formula(r, p, f)

    sb.section_header("L1 FFN: |x| via dual ReLU")
    sb.subsection_header("(MAX(0,x*10000) + MAX(0,-x*10000)) * 100")
    ffn1_rows = []
    for d in range(NUM_CANDIDATES):
        r = sb.label_row(f"ffn1[{d}]")
        ffn1_rows.append(r)
        for p in range(MAX_POS):
            x = cell(p, cand_rows[d])
            f = f'=(MAX(0,{x}*10000)+MAX(0,-{x}*10000))*100'
            sb.write_formula(r, p, f)

    return cand_rows, ffn1_rows


def build_final_h(sb, l1_resid_rows, ffn1_rows):
    """Final h = pad_to(h, 10) + FFN1 output.
    pad_to(h, 10) takes dims 0..9 of l1_resid. But the code does:
      h = pad_to(h, NUM_CANDIDATES)  # takes first 10 dims of 16-dim h
      h = h + ffn1_out
    So final_h[d] = l1_resid[d] + ffn1[d] for d=0..9
    But ffn1 only covers candidates (dims 5-14 originally → mapped to 0-9 of ffn1).
    Wait, let me re-read the code...

    candidates = h[..., CANDIDATES_START:CANDIDATES_START + NUM_CANDIDATES]  # dims 5..14
    ffn1_out = self.layer1_ffn(candidates)  # 10 outputs
    h = pad_to(h, NUM_CANDIDATES)  # h is now dims 0..9 of the 16-dim vector
    h = h + ffn1_out  # add ffn1 to dims 0..9

    So final_h[d] = l1_resid[d] + ffn1[d] for d=0..9
    where ffn1[d] = |candidates[d]| * stuff
    """
    sb.section_header("FINAL h (argmin input, 10 dims)")
    final_rows = []
    for d in range(NUM_CANDIDATES):
        r = sb.label_row(f"h[{d}]")
        final_rows.append(r)
        for p in range(MAX_POS):
            resid = cell(p, l1_resid_rows[d])
            ffn = cell(p, ffn1_rows[d])
            f = f'={resid}+{ffn}'
            sb.write_formula(r, p, f)
    return final_rows


def build_prediction(sb, final_rows, tok_row):
    """Build prediction: argmin of final_h at the last non-empty position."""
    sb.section_header("PREDICTION")

    # For each position, compute the min and argmin of the 10 final_h values
    sb.subsection_header("Min value at each position")
    min_row = sb.label_row("min")
    for p in range(MAX_POS):
        parts = ",".join(cell(p, final_rows[d]) for d in range(NUM_CANDIDATES))
        f = f'=MIN({parts})'
        sb.write_formula(min_row, p, f)

    sb.subsection_header("Predicted digit (argmin)")
    pred_row = sb.label_row("digit")
    for p in range(MAX_POS):
        # MATCH finds first match in array
        parts = [cell(p, final_rows[d]) for d in range(NUM_CANDIDATES)]
        min_c = cell(p, min_row)
        # Build a MATCH against the column of final_h values
        # We need a vertical range or use nested IFs...
        # Easiest: use a series of IF statements
        f = f'=IF({cell(p, tok_row)}="","",MATCH({min_c},{parts[0]}:{parts[-1]},0)-1)'
        sb.write_formula(pred_row, p, f)

    sb.subsection_header("Predicted token")
    token_row = sb.label_row("token")
    for p in range(MAX_POS):
        pred_c = cell(p, pred_row)
        f = f'=IF({pred_c}="","",INDEX(Ref!$A$2:$A$15,{pred_c}+1))'
        sb.write_formula(token_row, p, f)

    # Summary: prediction for last filled position
    sb.skip()
    r = sb.label_row("PREDICTION →")
    sb.ws.cell(r, 1).font = Font(bold=True, size=14)
    # Find the last non-empty token column
    seq_len_formula = f'COUNTA({cell(0, tok_row)}:{cell(MAX_POS-1, tok_row)})'
    # The predicted token at the last position
    # INDEX into the token_row using seq_len as column offset
    f = f'=IF({seq_len_formula}=0,"",INDEX({cell(0, token_row)}:{cell(MAX_POS-1, token_row)},1,{seq_len_formula}))'
    sb.write_formula(r, 0, f)
    sb.ws.cell(r, COL_OFFSET).font = Font(bold=True, size=14, color="FF0000")

    return min_row, pred_row, token_row


def apply_formatting(ws, tok_row, final_rows):
    """Apply column widths, freeze panes, and conditional formatting."""
    # Column A width
    ws.column_dimensions['A'].width = 16

    # Token columns width
    for p in range(MAX_POS):
        ws.column_dimensions[col(p)].width = 14

    # Freeze panes: keep rows 1-3 (header + token input + token ID) visible
    ws.freeze_panes = f'A{tok_row + 2}'


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TinyAdder4"

    sb = SheetBuilder(ws)

    # Build Ref sheet
    build_ref_sheet(wb)

    # Default tokens: <bos> 1 2 3 4 + 5 6 7 8 =
    default_tokens = ["<bos>", "1", "2", "3", "4", "+", "5", "6", "7", "8", "="]

    # Build sections
    tok_row, id_row = build_input_section(sb, default_tokens)
    emb_rows = build_embedding_section(sb, id_row)
    h3_out_row = build_l0_head3(sb, emb_rows)
    h4_out_row = build_l0_head4(sb, emb_rows)
    resid_rows = build_l0_attn_residual(sb, emb_rows, h3_out_row, h4_out_row)
    ffn_rows = build_l0_ffn(sb, resid_rows)
    wide_rows = build_widened_residual(sb, resid_rows, ffn_rows)
    l1_out_row = build_l1_attention(sb, wide_rows)
    l1_resid_rows = build_l1_residual(sb, wide_rows, l1_out_row)
    cand_rows, ffn1_rows = build_candidates_and_ffn1(sb, l1_resid_rows)
    final_rows = build_final_h(sb, l1_resid_rows, ffn1_rows)
    build_prediction(sb, final_rows, tok_row)

    apply_formatting(ws, tok_row, final_rows)

    # Save
    outpath = "tinyadder_4.xlsx"
    wb.save(outpath)
    print(f"Saved {outpath} ({sb.current_row()} rows)")


if __name__ == "__main__":
    main()
